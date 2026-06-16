"""Render docs/uat/UAT-cafe-pos.xlsx from uat_cases data.

Usage (repo root): uv run --with openpyxl python tools/generate_uat_workbook.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from uat_cases import EXTRA_MODULE, MODULES, TECH_MODULE

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = REPO_ROOT / "docs" / "uat" / "UAT-cafe-pos.xlsx"

HEADERS = ["รหัส", "สถานการณ์ทดสอบ", "ขั้นตอน", "ผลที่คาดหวัง", "ผล", "หมายเหตุ", "ผู้ทดสอบ", "วันที่"]
COL_WIDTHS = [14, 38, 48, 42, 10, 24, 16, 12]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def build_module_sheet(wb: Workbook, module: dict) -> str:
    ws = wb.create_sheet(module["sheet"][:31])
    ws["A1"] = f"การทดสอบ: {module['sheet']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"อ้างอิง: {module['tor']}"

    header_row = 4
    for col, (title, width) in enumerate(zip(HEADERS, COL_WIDTHS, strict=True), start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    dv = DataValidation(type="list", formula1='"ผ่าน,ไม่ผ่าน"', allow_blank=True)
    ws.add_data_validation(dv)

    for idx, case in enumerate(module["cases"], start=1):
        row = header_row + idx
        steps = "\n".join(f"{i}. {s}" for i, s in enumerate(case["steps"], start=1))
        values = [
            f"UAT-{module['prefix']}-{idx:02d}",
            case["scenario"],
            steps,
            case["expected"],
            "",
            "",
            "",
            "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
        dv.add(ws.cell(row=row, column=5))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return ws.title


def build_summary_sheet(wb: Workbook, module_sheets: list[tuple[str, int]]) -> None:
    ws = wb["สรุปผลและลงนาม"]
    ws["A1"] = "สรุปผลการทดสอบระบบ (UAT) — ระบบ POS คาเฟ่ตะวันอ้อมข้าว"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = (
        "เกณฑ์การตรวจรับ: รายการทดสอบในขอบเขต TOR (ข้อ 3.1–3.11 และข้อ 4) "
        "ต้องมีผล 'ผ่าน' ครบทุกรายการ รายการในชีต 'ฟังก์ชันเพิ่มเติม (นอก TOR)' "
        "ไม่อยู่ในเกณฑ์การตรวจรับ"
    )
    ws["A3"].alignment = WRAP
    ws.merge_cells("A3:E3")
    ws.row_dimensions[3].height = 45

    headers = ["โมดูล", "จำนวนข้อทดสอบ", "ผ่าน", "ไม่ผ่าน", "ยังไม่ทดสอบ"]
    header_row = 5
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    widths = [38, 16, 10, 10, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for idx, (sheet_name, case_count) in enumerate(module_sheets, start=1):
        row = header_row + idx
        passed = f"COUNTIF('{sheet_name}'!E:E,\"ผ่าน\")"
        failed = f"COUNTIF('{sheet_name}'!E:E,\"ไม่ผ่าน\")"
        ws.cell(row=row, column=1, value=sheet_name).border = BORDER
        ws.cell(row=row, column=2, value=case_count).border = BORDER
        ws.cell(row=row, column=3, value=f"={passed}").border = BORDER
        ws.cell(row=row, column=4, value=f"={failed}").border = BORDER
        ws.cell(row=row, column=5, value=f"={case_count}-{passed}-{failed}").border = BORDER

    sign_row = header_row + len(module_sheets) + 3
    ws.cell(row=sign_row, column=1, value="ผลการตรวจรับ:  [  ] ผ่านการตรวจรับ    [  ] ไม่ผ่านการตรวจรับ")
    for label, col in ((("ผู้ว่าจ้าง"), 1), (("ผู้รับจ้าง"), 4)):
        base = sign_row + 3
        ws.cell(row=base, column=col, value=f"ลงชื่อ ____________________ ({label})")
        ws.cell(row=base + 2, column=col, value="ชื่อ-สกุล ____________________")
        ws.cell(row=base + 4, column=col, value="วันที่ ____________________")


def main() -> None:
    wb = Workbook()
    wb.active.title = "สรุปผลและลงนาม"

    module_sheets: list[tuple[str, int]] = []
    for module in [*MODULES, TECH_MODULE, EXTRA_MODULE]:
        title = build_module_sheet(wb, module)
        if module is not EXTRA_MODULE:
            module_sheets.append((title, len(module["cases"])))

    build_summary_sheet(wb, module_sheets)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
